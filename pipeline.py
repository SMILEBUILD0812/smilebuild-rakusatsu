#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SmileBuild 落札情報パイプライン
================================
毎日 GitHub Actions で実行され、以下を行う：

  P1  関東地整／PPI から入札・契約結果を取得         -> fetch_results()
  P2  入札調書の明細から「入札参加者」を取得          -> fetch_bidders()
  P3  経審/許可データと突合して従業員数・所在地を補強  -> enrich()
  蓄積  全期間を history.json に貯める（履歴・年間集計の素）
  出力  ツール用の data.json を生成（会社集計はツール側で実施）
  通知  前回からの「新着の該当案件」をメール送信

【重要・キャリブレーション】
  役所サイトの実ファイル構造は独特で、中身を見ずに完璧に抜くのは不可能です。
  実データから値を抜く 3 つの関数（fetch_results / fetch_bidders / enrich）の
  中の "★CALIBRATE★" の箇所だけ、実ファイルを 1 度開いて合わせてください。
  それ以外（蓄積・集計・data.json生成・差分検知・メール）は完成しています。

  まず動作確認は:  python pipeline.py --sample      (ダミーで全工程を通す)
  実運用は:        python pipeline.py                (実取得)
"""

import os, sys, json, hashlib, smtplib, datetime, re, time
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

# ---- 任意依存（実取得時のみ必要。--sample では不要）----
try:
    import requests
    from bs4 import BeautifulSoup
except Exception:
    requests = None
    BeautifulSoup = None
import io
try:
    import xlrd            # 入札結果Excel(.xls)解析用
except Exception:
    xlrd = None
try:
    import openpyxl          # 入札結果Excel(.xlsx)解析用（東北など）
except Exception:
    openpyxl = None
import urllib.parse as _up

# =========================================================
#  設定（GitHub Secrets / 環境変数で上書き可）
# =========================================================
ROOT      = os.path.dirname(os.path.abspath(__file__))
HISTORY   = os.path.join(ROOT, "history.json")     # 全期間の蓄積
DATA_JSON = os.path.join(ROOT, "data.json")        # ツールが読む最新データ
STATE     = os.path.join(ROOT, "state.json")       # 前回通知済みID
SALES_STATUS_CSV = os.path.join(ROOT, "sales-status.csv")  # 落札くんから書き出した営業ステータス

# 通知の条件（＝弘晃のターゲット）。Secrets で調整可。
NOTIFY_MIN_YEN = int(os.environ.get("NOTIFY_MIN_YEN") or 80_000_000)    # 8千万
NOTIFY_MAX_YEN = int(os.environ.get("NOTIFY_MAX_YEN") or 250_000_000)   # 2.5億
NOTIFY_PREFS   = [p for p in os.environ.get("NOTIFY_PREFS", "").split(",") if p]  # 空=全部

# メール本文に載せる上位社数（多すぎると読まれない、少なすぎるとアクション数が足りない）
MAIL_TOP_N = int(os.environ.get("MAIL_TOP_N") or 30)

# 関東1都6県（ターゲット度のボーナス）
KANTO_PREFS = {"千葉県", "東京都", "神奈川県", "埼玉県", "茨城県", "栃木県", "群馬県"}
# 関東+隣接（second-tier ボーナス）
KANTO_NEIGHBOR = {"山梨県", "長野県", "新潟県", "福島県", "静岡県"}
# 工事系の主要工種（弘晃のスイートスポット）
SWEET_KINDS = ("一般土木", "維持修繕", "舗装", "河川", "道路", "土木")

# メール（GitHub Secrets に登録）
SMTP_HOST = os.environ.get("SMTP_HOST", "")
SMTP_PORT = int(os.environ.get("SMTP_PORT") or "587")
SMTP_USER = os.environ.get("SMTP_USER", "")
SMTP_PASS = os.environ.get("SMTP_PASS", "")
MAIL_FROM = os.environ.get("MAIL_FROM", SMTP_USER)
MAIL_TO   = os.environ.get("MAIL_TO", "")

HTTP_UA = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}

# ---- 全国の地方整備局：入札結果Excel（高知談合対応の全国共通様式）を自動取得 ----
# 各局「ページ →(必要なら年度サブページ)→ 工事Excel」。様式は全国共通なので同じパーサーで抜ける。
#   pat = 工事Excelのファイル名の特徴（業務=_g/gyoumu等は別途除外）。None=工事見出し配下を採用。
#   sub = 年度サブページを辿る場合のリンク条件（None=トップ直下のみ）。
BUREAU_SOURCES = [
    # ===== 取り込み済み（実データ検証済・直近2年で約3,000件）=====
    {"name": "関東地方整備局", "page": "https://www.ktr.mlit.go.jp/nyuusatu/nyuusatu00004729.html",
     "pat": None, "sub": None},
    {"name": "四国地方整備局", "page": "http://www.skr.mlit.go.jp/send/nyuusatu/index.html",
     "pat": r"kou?zi", "sub": None},
    {"name": "東北地方整備局", "page": "http://www.thr.mlit.go.jp/Bumon/B00013/K00730/nyusatu/nyusatukekka/index.html",
     "pat": r"kouji", "sub": r"kensetsu|令和|R\d"},
    {"name": "北陸地方整備局", "page": "https://www.hrr.mlit.go.jp/keiyaku/kekka.html",
     "pat": r"kouji_\d+\.xls", "sub": r"kekka/r\d+\.html"},
    {"name": "九州地方整備局", "page": "https://www.qsr.mlit.go.jp/nyusatu_joho/keiyaku/nyusatu_data/",
     "pat": None, "sub": r"R\d{2}_\d+\.html"},
    # ===== 取り込み不可（実調査の結果、現役のExcel公表が見つからず）=====
    #   ・近畿地方整備局：DataLink記載のn_info/nyusatukekka は平成28年で更新停止。
    #                    現役の置き場所が見つからず（PPI誘導に切替えた可能性）。
    #   ・中部地方整備局：contract/ 配下に Excel 公開なし。「入札の経過」は PPI に誘導。
    #                    transparency Excel は港湾空港部のみ（土木工事は対象外）。
    #   ・中国地方整備局：hattyu/result/index.htm はほぼ空ページ(444bytes)。Excel未公開。
    #   ・北海道開発局：DataLink記載URLは全て404。現役URL不明。
    #   ・沖縄総合事務局：国交省「入札契約データリンク集」にExcel公表のリンクなし。
    #   ※追加検討する場合の手掛かりとして、国交省「入札契約データリンク集」を再確認：
    #     https://www.mlit.go.jp/tec/nyuusatu/datalink.html
    #
    # ===== 都道府県（実調査の結論：取り込み不可）=====
    # 公共工事入札契約適正化法(第8条)で全都道府県に公表が義務付けられているが、
    # 公表"方法"が「電子調達ポータル(検索フォーム型)」に集約されており、整備局のような
    # 月次/年次のExcelを固定ページに置く形では公開していない。実調査:
    #   ・東京都    ✕ 東京都電子調達システム(e-procurement.metro.tokyo.lg.jp)
    #   ・神奈川県  ✕ かながわ電子入札共同システム
    #   ・千葉県    ✕ ちば電子調達システム
    #   ・埼玉県    ✕ 埼玉県電子入札共同システム
    #   ・茨城県    ✕ ppi.cals-ibaraki.lg.jp
    #   ・群馬県    ✕ 電子調達(トップで誘導検出)
    #   ・福岡県    ✕ 電子調達ポータル
    #   ・大阪府    ✕ 大阪府電子契約システム
    #   ※残り県も同じ「電子入札共同システム」系パターンに集約されており、Excel直リンク
    #     公開は事実上ゼロ。都道府県は本データ源では対象外（仕様としての限界）。
    # 都道府県・市町村まで網羅したい場合は、有料の入札情報サービス(NJSS等)の併用が現実解。
]
# 限定したい場合は Secrets/Variables BUREAUS_ONLY="関東,近畿" を設定（空=全部）
MAX_FILES_PER_BUREAU = int(os.environ.get("MAX_FILES_PER_BUREAU") or 8)

# ---- データ源：国交省 全地方整備局（工事のみ）----
# PPI(入札情報サービス)が全整備局を横断する公式の統合検索。工事専用の検索画面がある：
#   工事検索  https://www.i-ppi.jp/ippi/SearchServices/web/Koji/Kokoku/Search.aspx
# PPIは.aspxフォームのため取得には Playwright が必要（USE_PPI=1 で有効化）。
USE_PPI       = os.environ.get("USE_PPI", "0") == "1"
PPI_KOJI_URL  = "https://www.i-ppi.jp/ippi/SearchServices/web/Koji/Kokoku/Search.aspx"  # ★工事専用★
LOOKBACK_DAYS = int(os.environ.get("LOOKBACK_DAYS") or "45")   # 何日分さかのぼるか

# PPIを使わない場合：各地方整備局の入札結果ページを巡回。
# ★CALIBRATE★ 実URL/年度ディレクトリは年度替わりで変わるため、稼働時に最新を設定。
BUREAUS = [
    ("北海道開発局",     "https://www.hkd.mlit.go.jp/"),
    ("東北地方整備局",   "https://www.thr.mlit.go.jp/"),
    ("関東地方整備局",   "https://www.ktr.mlit.go.jp/nyuusatu/index00000028.html"),
    ("北陸地方整備局",   "https://www.hrr.mlit.go.jp/"),
    ("中部地方整備局",   "https://www.cbr.mlit.go.jp/"),
    ("近畿地方整備局",   "https://www.kkr.mlit.go.jp/"),
    ("中国地方整備局",   "https://www.cgr.mlit.go.jp/"),
    ("四国地方整備局",   "https://www.skr.mlit.go.jp/"),
    ("九州地方整備局",   "https://www.qsr.mlit.go.jp/"),
    ("沖縄総合事務局",   "https://www.dc.ogb.go.jp/"),
]
# 特定の整備局だけにしたい場合は Secrets BUREAUS_ONLY="関東,東北" を設定（空=全部）
BUREAUS_ONLY = [b for b in os.environ.get("BUREAUS_ONLY", "").split(",") if b]

TODAY = datetime.date.today().isoformat()

# ---- 工事のみ抽出（測量・建設コンサルタント等の"業務"を除外）----
_GYOMU_RE = re.compile(
    r"業務|委託|コンサル|設計(?!施工)|測量|調査|点検|診断|計画策定|検討|"
    r"監督支援|積算|技術審査|補償|用地|資料作成|データ作成|システム|"
    r"観測|解析|試験|研究|清掃|運転|運用|警備|賃貸|借[上り]|購入|物品|役務")
_KOJI_RE = re.compile(
    r"工事|工区|舗装|護岸|改良|拡幅|補修|補強|築堤|樋管|樋門|橋梁|橋りょう|"
    r"トンネル|ずい道|法面|のり面|電線共同溝|無電柱|区画線|防護柵|擁壁|"
    r"盛土|切土|浚渫|河道|耐震|塗装|遮音|排水|管渠|構築|新設|設置工")
def is_works(name="", kind="", category=""):
    """工事=True / 測量・建設コンサルタント等の業務=False。"""
    cat = str(category or "")
    if re.search(r"工事", cat) and not re.search(r"業務|コンサル|委託", cat):
        return True
    if re.search(r"業務|コンサル|委託", cat):
        return False
    blob = "%s %s %s" % (cat, kind, name)
    if re.search(r"設計施工|設計・施工|DB方式", blob):   # 設計施工一括は工事
        return True
    if _GYOMU_RE.search(blob):
        return False
    if _KOJI_RE.search(blob):
        return True
    return True   # 工事検索前提。判定不能は工事扱い（業務語があれば上で除外済み）

def log(*a):
    print("[%s]" % datetime.datetime.now().strftime("%H:%M:%S"), *a, flush=True)

def case_id(c):
    key = "%s|%s|%s|%s|%s" % (c.get("date"), c.get("name"), c.get("company"), c.get("org"), c.get("amount"))
    return hashlib.md5(key.encode("utf-8")).hexdigest()[:12]

def pref_of(s):
    if not s: return ""
    m = re.search(r"(東京都|北海道|(?:京都|大阪)府|(?:神奈川|和歌山|鹿児島)県|..県)", str(s))
    return m.group(1) if m else ""

def city_of(s):
    if not s: return ""
    p = pref_of(s); s = str(s)
    rest = s[s.find(p)+len(p):] if p else s
    m = re.match(r"\s*([^\d０-９]{1,10}?[市区町村郡])", rest)
    return m.group(1) if m else ""

# =========================================================
#  P1  入札・契約結果の取得
# =========================================================
def fetch_results(sample=False):
    """落札結果のリストを返す。各要素は data モデル(dict)。工事のみ。"""
    if sample:
        cases = _sample_cases()
    elif USE_PPI:
        cases = _fetch_ppi()
    else:
        cases = _fetch_ktr()      # ★本命：関東地整の入札結果Excelを自動取得・解析

    # 工事のみに限定（測量・建設コンサルタント等の業務を除外）
    before = len(cases)
    cases = [c for c in cases if is_works(c.get("name", ""), c.get("kind", ""), c.get("category", ""))]
    if before:
        log("工事フィルタ:", before, "→", len(cases), "件（業務・コンサルを除外）")
    log("P1 取得:", len(cases), "件")
    return cases

def _http_get(url):
    """(status, html文字列)。サーバのcharset未指定に備えUTF-8→cp932で復号。"""
    try:
        r = requests.get(url, timeout=50, headers={**HTTP_UA, "Referer": url})
    except Exception as e:
        log("  取得エラー", url, e); return 0, ""
    raw = r.content
    for enc in ("utf-8", "cp932", "euc-jp"):
        try: return r.status_code, raw.decode(enc)
        except Exception: pass
    return r.status_code, ""

_NG_FILE = re.compile(r"gyoumu|業務|youshiki|様式|buppin|物品|zuikei|随契|teinyu|低入|consul|コンサル|manual|tebiki", re.I)

def _excel_links(base, html, pat):
    out = []
    for m in re.findall(r'href="([^"]+\.(?:xls|xlsx))"', html or "", re.I):
        f = m.split("/")[-1]
        if _NG_FILE.search(f):           continue       # 業務・様式・物品などは除外
        if pat and not re.search(pat, f, re.I): continue
        out.append(_up.urljoin(base, m))
    return list(dict.fromkeys(out))

def _koji_links_by_heading(base, html):
    """関東のように「令和X年度 工事」見出しの配下に並ぶExcelを拾う（業務見出しは除外）。"""
    parts = re.split(r"(令和\s*\d+\s*年度[　\s]*(?:工事|業務))", html or "")
    out = []
    for i in range(1, len(parts), 2):
        if "業務" in parts[i]:
            continue
        body = parts[i + 1] if i + 1 < len(parts) else ""
        for m in re.findall(r'href="([^"]+\.(?:xls|xlsx))"', body, re.I):
            if not _NG_FILE.search(m.split("/")[-1]):
                out.append(_up.urljoin(base, m))
    return list(dict.fromkeys(out))

def _sub_links(base, html, pat):
    """年度サブページ等のリンク（hrefがpatに合致）。新しい順に。"""
    out = [ _up.urljoin(base, a) for a in re.findall(r'href="([^"]+\.html?[^"]*)"', html or "", re.I)
            if re.search(pat, a, re.I) ]
    return sorted(dict.fromkeys(out), reverse=True)

def _locate_koji_xls(cfg):
    st, html = _http_get(cfg["page"])
    if not html:
        return []
    urls = _koji_links_by_heading(cfg["page"], html)        # 関東型（見出し配下）
    if not urls:
        urls = _excel_links(cfg["page"], html, cfg.get("pat"))  # 四国型（トップ直下＋ファイル名）
    if not urls and cfg.get("sub"):                          # 近畿/東北型（年度サブページ）
        for sp in _sub_links(cfg["page"], html, cfg["sub"])[:4]:
            _, h2 = _http_get(sp)
            urls += _excel_links(sp, h2, cfg.get("pat"))
    urls = list(dict.fromkeys(urls))
    urls.sort(reverse=True)                                  # ファイル名降順≒新しい年月を優先
    return urls[:MAX_FILES_PER_BUREAU]

def _read_excel_rows(blob, _hint=None):
    head = blob[:8]
    if head[:2] == b"PK":                       # 本物の .xlsx（ZIP）
        if openpyxl is None: raise RuntimeError("openpyxl未導入")
        wb = openpyxl.load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
        sh = wb[wb.sheetnames[0]]
        return [list(r) for r in sh.iter_rows(values_only=True)]
    if head[:4] == b"\xD0\xCF\x11\xE0":          # OLE2 = .xls（拡張子が.xlsxでも中身がこれの場合あり）
        if xlrd is None: raise RuntimeError("xlrd未導入")
        wb = xlrd.open_workbook(file_contents=blob); sh = wb.sheet_by_index(0)
        return [[sh.cell_value(r, c) for c in range(sh.ncols)] for r in range(sh.nrows)]
    raise RuntimeError("Excelでない応答（HTML/エラー等）")

def _fetch_ktr():
    """全国の地方整備局の入札結果Excel（全国共通様式）を自動取得・解析。
       BUREAU_SOURCES に列挙した局を巡回。ブラウザ操作不要。"""
    if requests is None:
        log("requests が無いため取得スキップ。`pip install -r requirements.txt`"); return []
    if xlrd is None and openpyxl is None:
        log("xlrd/openpyxl が無いため取得スキップ。`pip install xlrd openpyxl`"); return []
    targets = [c for c in BUREAU_SOURCES
               if (not BUREAUS_ONLY or any(k in c["name"] for k in BUREAUS_ONLY))]
    log("対象整備局:", ", ".join(c["name"] for c in targets) or "（なし）")
    cases = []
    for cfg in targets:
        try:
            urls = _locate_koji_xls(cfg)
        except Exception as e:
            log("  %s 置き場所探索エラー: %s" % (cfg["name"], e)); urls = []
        if not urls:
            log("  %s … 工事Excelを検出できず（要URL確認）" % cfg["name"]); continue
        n_b = len(cases)
        for u in urls:
            try:
                blob = requests.get(u, timeout=60, headers={**HTTP_UA, "Referer": cfg["page"]}).content
                rows = _read_excel_rows(blob)
                cases += _parse_result_xls(rows, cfg["name"], u)
                time.sleep(0.8)  # 礼儀
            except Exception as e:
                log("    Excel取得/解析エラー %s: %s" % (u.split("/")[-1], e))
        log("  %s … %d件（%dファイル）" % (cfg["name"], len(cases) - n_b, len(urls)))
    return cases

def _xv_S(v):
    """セル値→文字列。"""
    if v is None: return ""
    if isinstance(v, float): return str(int(v)) if v == int(v) else str(v)
    if isinstance(v, (datetime.datetime, datetime.date)):
        return (v.date() if isinstance(v, datetime.datetime) else v).isoformat()
    return str(v).strip()

def _xv_D(v):
    """セル値→YYYY-MM-DD。datetime/Excel日付シリアル/YYYY/MM/DD/和暦に対応。"""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return (v.date() if isinstance(v, datetime.datetime) else v).isoformat()
    if isinstance(v, (int, float)) and v > 20000:
        try: return xlrd.xldate.xldate_as_datetime(v, 0).date().isoformat()
        except Exception: return ""
    s = _xv_S(v)
    m = re.search(r"(\d{4})[/.\-年](\d{1,2})[/.\-月](\d{1,2})", s)
    if m: return "%04d-%02d-%02d" % (int(m.group(1)), int(m.group(2)), int(m.group(3)))
    m = re.search(r"(令和|平成|R|H)\s*(\d{1,2})[/.\-年]\s*(\d{1,2})[/.\-月]\s*(\d{1,2})", s)
    if m:
        base = 2018 if m.group(1) in ("令和", "R") else 1988
        return "%04d-%02d-%02d" % (base + int(m.group(2)), int(m.group(3)), int(m.group(4)))
    return ""

def _xv_N(v):
    """セル値→float。文字列中の数字も拾う。"""
    if isinstance(v, (int, float)): return float(v)
    t = re.sub(r"[^\d.]", "", _xv_S(v)); return float(t) if t else None

def _parse_result_xls(rows, bureau, source_url):
    """入札結果Excelを解析（フォーマット自動判定）：
       A. per-bidder型（1業者=1行の調書）… 国交省共通様式（関東/東北/北陸/四国/九州 等）
       B. flat型（1工事=1行）… 適正化法に基づく「公共工事の契約状況の公表」（都道府県・市町村に多い）
       どちらでもなければ空。"""
    # ヘッダー行を見つけて様式判定
    H = None; HDR = None; mode = None
    for r in range(min(20, len(rows))):
        line = "|".join(_xv_S(x) for x in rows[r])
        has_name = ("工事名" in line or "件名" in line or "公共工事の名称" in line)
        if not has_name: continue
        if "入札業者名" in line or "業者名" in line:
            H, HDR, mode = r, [_xv_S(x) for x in rows[r]], "per_bidder"; break
        if any(k in line for k in ("契約の相手方", "受注者", "落札者", "商号", "契約者")):
            H, HDR, mode = r, [_xv_S(x) for x in rows[r]], "flat"; break
    if H is None: return []
    def col(*keys):
        for c, x in enumerate(HDR):
            if any(k in x for k in keys): return c
        return None
    if mode == "per_bidder":
        return _parse_per_bidder(rows, HDR, H, col, bureau, source_url)
    return _parse_flat(rows, HDR, H, col, bureau, source_url)

def _parse_per_bidder(rows, HDR, H, col, bureau, source_url):
    """A. 1業者=1行の調書型。工事単位にグループ化、備考='落札'で落札者を特定。"""
    ci = dict(org=col("部局名", "発注機関"), name=col("工事名", "件名"),
              bid=col("入札年月日", "開札年月日", "入札日", "開札日"),
              ctr=col("契約年月日", "契約日"), kind=col("工種"),
              company=col("入札業者名", "業者名"),
              addr=col("本店住所", "本社", "住所", "所在地"),
              yotei=col("予定価格"), memo=col("備考"), result=col("入札結果"))
    if ci["company"] is None or ci["name"] is None or ci["result"] is None:
        return []
    rc = ci["result"]; amt_cols = [rc, rc + 2, rc + 4]
    memo_c = ci["memo"] if ci["memo"] is not None else (len(HDR) - 1)
    groups = {}; order = []
    for r in range(H + 1, len(rows)):
        row = rows[r]
        def cell(c): return row[c] if (c is not None and c < len(row)) else ""
        comp = _xv_S(cell(ci["company"])); name = _xv_S(cell(ci["name"]))
        if not comp or not name: continue
        key = (_xv_S(cell(ci["org"])), name, _xv_D(cell(ci["bid"])))
        if key not in groups:
            groups[key] = {"org": _xv_S(cell(ci["org"])), "name": name, "date": _xv_D(cell(ci["bid"])),
                           "ctr": _xv_D(cell(ci["ctr"])), "kind": _xv_S(cell(ci["kind"])),
                           "yotei": _xv_N(cell(ci["yotei"])), "bidders": [],
                           "winner": None, "win_amt": None, "win_addr": ""}
            order.append(key)
        g = groups[key]
        amt = None
        for c in amt_cols:
            n = _xv_N(cell(c))
            if n is not None and n > 0: amt = n
        g["bidders"].append({"company": comp, "pref": pref_of(_xv_S(cell(ci["addr"]))), "amount": amt})
        if "落札" in _xv_S(cell(memo_c)) or "落札" in _xv_S(cell(rc)):
            g["winner"] = comp; g["win_amt"] = amt
            g["win_addr"] = _xv_S(cell(ci["addr"]))
    out = []
    for key in order:
        g = groups[key]
        if not g["winner"]: continue
        rec = _blank_rec()
        rate = round(g["win_amt"] / g["yotei"] * 100, 2) if (g["win_amt"] and g["yotei"]) else None
        rec.update({
            "date": g["ctr"] or g["date"], "name": g["name"], "org": g["org"],
            "company": g["winner"], "amount": g["win_amt"], "rate": rate,
            "kind": g["kind"] or classify_kind(g["name"]),
            "pref": pref_of(g["win_addr"]), "city": city_of(g["win_addr"]),
            "category": "工事", "bureau": bureau,
            "bidders": [b for b in g["bidders"] if b["company"]],
            "docs": [{"label": "入札結果(%s)" % bureau, "url": source_url}],
        })
        out.append(rec)
    return out

def _parse_flat(rows, HDR, H, col, bureau, source_url):
    """B. 1工事=1行のフラット型（適正化法『公共工事の契約状況の公表』形式）。
       契約の相手方・契約金額・予定価格・落札率が1行に並ぶ。都道府県・市町村で多い。"""
    ci = dict(
        name=col("公共工事の名称", "工事名", "件名"),
        date=col("契約を締結", "契約年月日", "契約日", "落札日", "入札年月日", "入札日", "開札日"),
        org=col("発注者", "発注機関", "部局名", "契約担当官", "所管", "所属"),
        kind=col("工事種別", "工種"),
        company=col("契約の相手方", "受注者", "落札者", "商号", "契約者", "業者名"),
        addr=col("契約の相手方の住所", "相手方の住所", "受注者の住所", "本店住所", "本社", "住所", "所在地"),
        yotei=col("予定価格"),
        amount=col("当初契約額", "契約金額", "契約額", "落札金額", "落札価格"),
        rate=col("落札率"),
    )
    if ci["name"] is None or ci["company"] is None: return []
    out = []
    for r in range(H + 1, len(rows)):
        row = rows[r]
        def cell(c): return row[c] if (c is not None and c < len(row)) else ""
        name = _xv_S(cell(ci["name"])); comp = _xv_S(cell(ci["company"]))
        if not name or not comp: continue
        kind = _xv_S(cell(ci["kind"]))
        if kind and any(k in kind for k in ("業務", "コンサルタント", "コンサル")):
            continue                                      # 行レベルで業務を除外
        amt = _xv_N(cell(ci["amount"]))
        yt  = _xv_N(cell(ci["yotei"]))
        rv  = _xv_N(cell(ci["rate"]))
        if rv is not None and rv <= 1.5: rv *= 100        # 0.9 → 90 へ正規化
        rate = round(rv, 2) if rv else (round(amt / yt * 100, 2) if amt and yt else None)
        addr = _xv_S(cell(ci["addr"]))
        org  = _xv_S(cell(ci["org"])) or bureau
        rec = _blank_rec()
        rec.update({
            "date": _xv_D(cell(ci["date"])),
            "name": name, "org": org,
            "company": comp, "amount": amt, "rate": rate,
            "kind": kind or classify_kind(name),
            "pref": pref_of(addr), "city": city_of(addr),
            "category": "工事", "bureau": bureau,
            "bidders": [{"company": comp, "pref": pref_of(addr), "amount": amt}],
            "docs": [{"label": "入札結果(%s)" % bureau, "url": source_url}],
        })
        out.append(rec)
    return out

def _fetch_bureaus():
    """全地方整備局の入札結果ページを巡回して落札（工事）を集める（参考実装）。"""
    if requests is None:
        log("requests/bs4 が無いため取得スキップ。`pip install -r requirements.txt`")
        return []
    targets = [(n, u) for (n, u) in BUREAUS if (not BUREAUS_ONLY or any(k in n for k in BUREAUS_ONLY))]
    log("巡回対象:", len(targets), "整備局",
        ("（"+",".join(BUREAUS_ONLY)+"に限定）") if BUREAUS_ONLY else "（全部）")
    cases = []
    for bureau, url in targets:
        try:
            html = requests.get(url, timeout=30,
                                headers={"User-Agent": "SmileBuild-bot/1.0 (info@smile-build.com)"}).text
            soup = BeautifulSoup(html, "html.parser")
            # ★CALIBRATE★ 各整備局の入札結果テーブル（または結果Excelへのリンク）に合わせて調整。
            #   多くは「工事」「業務」が別ページ/別ファイルで公表されるため、工事側のみを辿るのが理想。
            #   どうしても混在する場合も、後段の is_works() で業務は除外される。
            n0 = len(cases)
            for tr in soup.select("table tr"):
                tds = [td.get_text(strip=True) for td in tr.select("td")]
                if len(tds) < 4:
                    continue
                rec = _blank_rec()
                rec.update({
                    "date":    _norm_date(tds[0]),
                    "name":    tds[1],
                    "org":     (tds[2] if len(tds) > 2 and tds[2] else bureau),
                    "company": tds[3] if len(tds) > 3 else "",
                    "amount":  _norm_yen(tds[4]) if len(tds) > 4 else None,
                    "rate":    _norm_num(tds[5]) if len(tds) > 5 else None,
                    "bureau":  bureau,
                })
                if rec["company"] and rec["name"]:
                    rec["kind"] = classify_kind(rec["name"])
                    cases.append(rec)
            log("  %s: %d件" % (bureau, len(cases) - n0))
            time.sleep(1.0)  # 礼儀
        except Exception as e:
            log("  %s 取得エラー: %s" % (bureau, e))
    return cases

def _fetch_ppi():
    """PPI(入札情報サービス)の『工事』検索で全整備局横断に取得（Playwright）。
       ★CALIBRATE★ フォーム項目（発注機関=全整備局/開札日範囲/工事種別）と
       結果テーブルのセレクタを、実画面に合わせて1度だけ調整する。"""
    try:
        from playwright.sync_api import sync_playwright
    except Exception:
        log("Playwright未導入。`pip install playwright && playwright install chromium`。USE_PPI=0で各整備局巡回に切替可")
        return []
    cases = []
    since = (datetime.date.today() - datetime.timedelta(days=LOOKBACK_DAYS))
    try:
        with sync_playwright() as p:
            br = p.chromium.launch(headless=True)
            pg = br.new_page()
            pg.goto(PPI_KOJI_URL, timeout=60000)   # ★工事専用URL（業務は別画面なので源で除外）
            # ★CALIBRATE★ ここで検索条件を入力：
            #   - 発注機関：国土交通省（全地方整備局）にチェック/選択
            #   - 開札日 from = since.isoformat()  / to = TODAY
            #   - 区分：落札結果
            #   例) pg.fill("#kaisatsu_from", since.strftime("%Y/%m/%d"))
            #       pg.click("#searchButton")
            #   結果テーブルを行ごとに読み、必要なら次ページへ。
            #   for row in pg.query_selector_all("table#result tr"): ...
            #       rec = _blank_rec(); rec.update({...}); cases.append(rec)
            br.close()
    except Exception as e:
        log("PPI取得エラー:", e)
    return cases

def _blank_rec():
    return {
        "date": "", "name": "", "org": "", "company": "",
        "amount": None, "rate": None, "kind": "", "period": "",
        "pref": "", "city": "", "employees": None,
        "category": "", "bureau": "",
        "docs": [], "bidders": [], "detail_url": "",
    }

KIND_RULES = [
    ("舗装", ["舗装"]), ("河川改修", ["河川", "築堤", "護岸", "樋管", "樋門"]),
    ("トンネル", ["トンネル", "ずい道"]), ("橋梁補修", ["橋", "橋梁"]),
    ("電線共同溝", ["電線共同溝", "無電柱"]), ("法面工", ["法面", "のり面"]),
    ("道路改良", ["道路", "改良", "拡幅"]), ("維持修繕", ["維持", "修繕", "補修"]),
]
def classify_kind(name):
    for label, keys in KIND_RULES:
        if any(k in (name or "") for k in keys):
            return label
    return "その他"

# =========================================================
#  P2  入札調書の明細 → 入札参加者
# =========================================================
def fetch_bidders(case):
    """その案件に応札した会社のリスト [{company, pref, amount}] を返す。"""
    url = case.get("detail_url")
    if not url or requests is None:
        return []
    bidders = []
    try:
        html = requests.get(url, timeout=30,
                            headers={"User-Agent": "SmileBuild-bot/1.0"}).text
        soup = BeautifulSoup(html, "html.parser")
        # ★CALIBRATE★ 入札調書の「参加業者・入札額」の表に合わせて取り出す
        for tr in soup.select("table tr"):
            tds = [td.get_text(strip=True) for td in tr.select("td")]
            if len(tds) < 2:
                continue
            name = tds[0]
            amt  = _norm_yen(tds[1])
            if name and not re.search(r"業者名|商号|会社", name):
                bidders.append({"company": name, "pref": pref_of(name), "amount": amt})
        time.sleep(1.0)  # 礼儀: 1秒空ける
    except Exception as e:
        log("fetch_bidders エラー:", e)
    return bidders

# =========================================================
#  P3  経審/許可データ突合 → 従業員数・所在地
# =========================================================
_KEISHIN = None
def _load_keishin():
    """経審/建設業許可の会社マスタを {正規化会社名: {pref, employees}} で返す。
       ★CALIBRATE★ 実データ(経審結果CSV等)を keishin.csv として置き、列を合わせる。"""
    global _KEISHIN
    if _KEISHIN is not None:
        return _KEISHIN
    _KEISHIN = {}
    path = os.path.join(ROOT, "keishin.csv")
    if not os.path.exists(path):
        log("keishin.csv 無し → 従業員数/所在地の突合はスキップ（任意）")
        return _KEISHIN
    import csv
    with open(path, encoding="utf-8") as f:
        for row in csv.DictReader(f):
            name = _normalize_co(row.get("商号") or row.get("会社名") or "")
            if not name:
                continue
            emp = row.get("職員数") or row.get("従業員数") or ""
            addr = row.get("所在地") or row.get("住所") or ""
            _KEISHIN[name] = {
                "pref": pref_of(addr),
                "city": city_of(addr),
                "employees": int(re.sub(r"\D", "", emp)) if re.search(r"\d", emp) else None,
            }
    log("経審マスタ:", len(_KEISHIN), "社")
    return _KEISHIN

def _normalize_co(s):
    # 会社名の表記ゆれ吸収（突合用）。完全一致は難しいので取りこぼしは出る。
    s = str(s or "")
    s = re.sub(r"[（(].*?[）)]", "", s)
    s = s.replace("株式会社", "").replace("有限会社", "").replace("(株)", "").replace("（株）", "")
    return re.sub(r"\s|　", "", s)

# ---- gBizINFO（経済産業省・無料API）で 所在地＋従業員数 を取得 ----
# 無料トークンを https://info.gbiz.go.jp/ で発行し、Secrets/環境変数 GBIZ_TOKEN に設定。
GBIZ_TOKEN = os.environ.get("GBIZ_TOKEN", "")
def _gbiz_lookup(name):
    if not (GBIZ_TOKEN and requests and name):
        return None
    try:
        r = requests.get("https://info.gbiz.go.jp/hojin/v1/hojin",
                         params={"name": name, "limit": 1},
                         headers={"X-hojinInfo-api-token": GBIZ_TOKEN, "Accept": "application/json"},
                         timeout=20)
        if r.status_code != 200:
            return None
        infos = (r.json() or {}).get("hojin-infos") or []
        if not infos:
            return None
        h = infos[0]
        loc = h.get("location") or ""
        emp = h.get("employee_number")
        return {"pref": pref_of(loc), "city": city_of(loc),
                "employees": int(emp) if (emp not in (None, "")) else None}
    except Exception as e:
        log("gBiz lookup失敗:", name, e)
        return None

def enrich(cases):
    """所在地（県・市）と従業員数を自動補完。
       1) gBizINFO API（GBIZ_TOKEN があれば）  2) keishin.csv（あれば）の順で照合。"""
    ks = _load_keishin()
    cache = {}
    def lookup(name):
        key = _normalize_co(name)
        if not key:
            return None
        if key in cache:
            return cache[key]
        info = _gbiz_lookup(name)
        if GBIZ_TOKEN:
            time.sleep(0.3)   # APIへの礼儀（レート制限回避）
        if not info and ks:
            info = ks.get(key)
        cache[key] = info
        return info
    hit = 0
    for c in cases:
        info = lookup(c.get("company"))
        if info:
            if not c.get("pref") and info.get("pref"): c["pref"] = info["pref"]
            if not c.get("city") and info.get("city"): c["city"] = info["city"]
            if c.get("employees") is None and info.get("employees") is not None: c["employees"] = info["employees"]
            hit += 1
        for b in c.get("bidders", []):
            bi = lookup(b.get("company"))
            if bi:
                b["pref"] = b.get("pref") or bi.get("pref", "")
                b["city"] = b.get("city") or bi.get("city", "")
                b["employees"] = bi.get("employees")
    log("所在地・従業員 補完:", hit, "/", len(cases),
        "（gBizトークン:", "有" if GBIZ_TOKEN else "無", "/ 経審マスタ:", len(ks), "社）")

# =========================================================
#  蓄積・出力・通知
# =========================================================
def load_json(path, default):
    if os.path.exists(path):
        try:
            with open(path, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return default

def save_json(path, obj):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False, indent=1)

def merge_history(history_cases, new_cases):
    by_id = {c["id"]: c for c in history_cases}
    added = []
    for c in new_cases:
        c["id"] = case_id(c)
        if c["id"] not in by_id:
            by_id[c["id"]] = c
            added.append(c)
        else:
            by_id[c["id"]].update({k: v for k, v in c.items() if v})  # 後勝ちで補強
    merged = sorted(by_id.values(), key=lambda x: x.get("date") or "", reverse=True)
    return merged, added

def _load_sales_status():
    """落札くんから書き出された sales-status.csv を読込み、{会社名(正規化)→ステータス文字列}を返す。
       ファイルが無い場合は空dict（=全社未着手扱い）。"""
    import csv
    if not os.path.exists(SALES_STATUS_CSV):
        return {}
    status_map = {}
    try:
        with open(SALES_STATUS_CSV, encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                name = (row.get("会社名") or row.get("company") or "").strip()
                stat = (row.get("営業ステータス") or row.get("status") or "").strip()
                if name and stat:
                    status_map[_normalize_co(name)] = stat
    except Exception as e:
        log("sales-status.csv 読込失敗（無視して継続）:", e)
    return status_map

def _company_lookup_counts(merged):
    """蓄積全件から、同社の累計落札件数を集計（ターゲット度の判定材料）"""
    counts = {}
    for c in merged:
        nm = _normalize_co(c.get("company") or "")
        if nm:
            counts[nm] = counts.get(nm, 0) + 1
    return counts

def target_score(c, company_counts=None):
    """1案件のターゲット度を0〜20点で算出。営業優先順位に使う。
       内訳：
         金額帯（1〜2億=+5、8千万-1億 or 2億-2.5億=+3、その他=+0）
         県（関東1都6県=+5、隣接=+3）
         工種（一般土木/維持修繕/舗装/河川/道路=+3）
         落札率（80-95%=+2、99%超=-3、80%未満=-3）
         同社累計落札件数（3件以上=+4、2件=+2）
         従業員（5〜25人=+3、補完なし=+1） ※gBiz補完済を想定
    """
    score = 0
    a = c.get("amount") or 0
    if 100_000_000 <= a <= 200_000_000:        score += 5
    elif 80_000_000 <= a < 100_000_000:         score += 3
    elif 200_000_000 < a <= 250_000_000:        score += 3
    pref = c.get("pref") or ""
    if pref in KANTO_PREFS:                     score += 5
    elif pref in KANTO_NEIGHBOR:                score += 3
    kind = c.get("kind") or ""
    if any(k in kind for k in SWEET_KINDS):     score += 3
    r = c.get("rate")
    if r is not None:
        if 80 <= r <= 95:                       score += 2
        elif r > 99 or r < 80:                  score -= 3
    if company_counts is not None:
        n = company_counts.get(_normalize_co(c.get("company") or ""), 1)
        if n >= 3:                              score += 4
        elif n >= 2:                            score += 2
    emp = c.get("employees")
    if emp is not None and 5 <= emp <= 25:      score += 3
    elif emp is None:                           score += 1     # 突合できてない=判定不能だが除外しない
    return score

def matches_target(c):
    """旧仕様の基本フィルタ（金額帯・県）。下流の優先順位は target_score で決める。"""
    a = c.get("amount")
    if a is None or not (NOTIFY_MIN_YEN <= a <= NOTIFY_MAX_YEN):
        return False
    if NOTIFY_PREFS and c.get("pref") and c["pref"] not in NOTIFY_PREFS:
        return False
    return True

# 営業ステータスの分類（落札くん側と整合）
# 「未着手」「対象外」以外（連絡済/手紙送付済/FAX送付済/商談中/契約/失注）はメール対象から除外
_EXCLUDE_STATUSES = {"連絡済", "手紙送付済", "FAX送付済", "商談中", "契約", "失注"}
def _is_actionable(c, status_map):
    """メール（手紙・FAX対応の対象）として有効か。営業ステータスが上記の除外集合に
       入っていれば False（既にアクション中・終了済）。"""
    if not status_map:
        return True
    nm = _normalize_co(c.get("company") or "")
    st = status_map.get(nm, "")
    return st not in _EXCLUDE_STATUSES

def _yen_jp(amount):
    """円→「1億2,345万円」形式に変換（メール本文の可読性向上）"""
    if amount is None: return "—"
    a = int(amount)
    if a >= 100_000_000:
        oku = a // 100_000_000
        man = (a % 100_000_000) // 10_000
        if man:  return "%d億%s万円" % (oku, "{:,}".format(man))
        return "%d億円" % oku
    return "{:,}万円".format(a // 10_000)

def _stars(score):
    """ターゲット度（0〜20点）を★1〜5に変換（メール表示用）"""
    if score >= 16: return "★★★★★"
    if score >= 13: return "★★★★☆"
    if score >= 10: return "★★★☆☆"
    if score >=  7: return "★★☆☆☆"
    if score >=  4: return "★☆☆☆☆"
    return "☆☆☆☆☆"

def send_email(top_targets, total_new, status_excluded_n=0):
    """ターゲット度上位の未着手会社のみをリッチ表でメール送信。
       top_targets はすでに target_score 降順でソート済み・上位N件。"""
    if not (SMTP_HOST and MAIL_TO):
        log("メール設定が無いため送信スキップ（候補 %d件）" % len(top_targets))
        return
    if not top_targets:
        log("候補ゼロ → メール送らず")
        return
    rows = ""
    for i, c in enumerate(top_targets, 1):
        bg = "#FAFCFA" if i % 2 == 0 else "#FFFFFF"
        rows += (
            "<tr style='background:%s'>"
            "<td style='padding:8px 10px;border-bottom:1px solid #eee;text-align:center;color:#999'>%d</td>"
            "<td style='padding:8px 10px;border-bottom:1px solid #eee;color:#C0392B;letter-spacing:-1px'>%s</td>"
            "<td style='padding:8px 10px;border-bottom:1px solid #eee;font-weight:700;color:#1b2e3c'>%s</td>"
            "<td style='padding:8px 10px;border-bottom:1px solid #eee;color:#5a6a72;font-size:12px'>%s</td>"
            "<td style='padding:8px 10px;border-bottom:1px solid #eee;color:#5a6a72;font-size:12px'>%s</td>"
            "<td style='padding:8px 10px;border-bottom:1px solid #eee'>%s</td>"
            "<td style='padding:8px 10px;border-bottom:1px solid #eee;text-align:right;font-weight:600'>%s</td>"
            "<td style='padding:8px 10px;border-bottom:1px solid #eee;text-align:right;color:#5a6a72'>%s%%</td>"
            "<td style='padding:8px 10px;border-bottom:1px solid #eee;text-align:right;color:#5a6a72'>%s</td>"
            "<td style='padding:8px 10px;border-bottom:1px solid #eee;text-align:center;color:#2d7a4f;font-weight:700'>%s社目</td>"
            "</tr>"
        ) % (
            bg, i, c.get("_stars", ""),
            c.get("company", "")[:18],
            (c.get("pref","") + (c.get("city","") or ""))[:14] or "—",
            (c.get("kind","") or "")[:12],
            (c.get("name","") or "")[:34],
            _yen_jp(c.get("amount")),
            "{:.1f}".format(c.get("rate") or 0) if c.get("rate") else "—",
            ("%d人" % c.get("employees")) if c.get("employees") else "—",
            c.get("_company_count", 1),
        )

    body = """
<div style="font-family:'Hiragino Kaku Gothic ProN','Yu Gothic',sans-serif;color:#1b2e3c;max-width:1100px">
  <h2 style="color:#2d7a4f;margin:0 0 4px;font-size:20px">本日の優先ターゲット {n} 社</h2>
  <p style="color:#5a6a72;font-size:12px;margin:0 0 8px;line-height:1.6">
    本日の取得 {tot} 件のうち、ターゲット度が高く<b>未着手</b>の会社のみ抜粋（上位 {n} 社／対象外・商談中・契約済を除外{exc}）。<br>
    並び順：<b>ターゲット度＞落札金額帯＞関東優先</b>。表示の★は県・金額帯・工種・落札率・累計落札・従業員規模を合成したスコア。
  </p>
  <p style="color:#5a6a72;font-size:12px;margin:0 0 14px">
    アクション目安：<b>★★★★★ → 即手紙（個別カスタム）</b>　<b>★★★★☆ → 手紙 or FAX</b>　<b>★★★☆☆ → FAX</b>
  </p>
  <table style="border-collapse:collapse;font-size:12px;width:100%;border:1px solid #ccc">
    <thead>
      <tr style="background:#1b2e3c;color:#fff">
        <th style="padding:8px 6px;text-align:center;width:30px">#</th>
        <th style="padding:8px 6px;text-align:center;width:90px">ターゲット度</th>
        <th style="padding:8px 10px;text-align:left">落札者</th>
        <th style="padding:8px 6px;text-align:left">所在地</th>
        <th style="padding:8px 6px;text-align:left">工種</th>
        <th style="padding:8px 10px;text-align:left">直近工事名</th>
        <th style="padding:8px 6px;text-align:right">落札金額</th>
        <th style="padding:8px 6px;text-align:right">落札率</th>
        <th style="padding:8px 6px;text-align:right">従業員</th>
        <th style="padding:8px 6px;text-align:center">累計</th>
      </tr>
    </thead>
    <tbody>{rows}</tbody>
  </table>
  <p style="color:#5a6a72;font-size:11px;margin-top:10px;line-height:1.6">
    詳細・履歴・参加業者ランキング・営業ステータスの更新は<br>
    <a href="https://smilebuild0812.github.io/smilebuild-rakusatsu/" style="color:#2d7a4f">→ 落札くん（Pages版）</a><br>
    で「★優先ターゲット」ボタンを押すと、同条件で他の候補も全件閲覧できます。
  </p>
</div>""".format(
    n=len(top_targets),
    tot=total_new,
    rows=rows,
    exc=("（営業中除外 %d社）" % status_excluded_n) if status_excluded_n else ""
)
    msg = MIMEMultipart("alternative")
    msg["Subject"] = "【SmileBuild】本日の優先ターゲット %d社（%s）" % (len(top_targets), TODAY)
    msg["From"], msg["To"] = MAIL_FROM, MAIL_TO
    msg.attach(MIMEText(body, "html", "utf-8"))
    try:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=30) as s:
            s.starttls()
            if SMTP_USER:
                s.login(SMTP_USER, SMTP_PASS)
            s.sendmail(MAIL_FROM, [a.strip() for a in MAIL_TO.split(",")], msg.as_string())
        log("メール送信:", len(top_targets), "社 →", MAIL_TO)
    except Exception as e:
        log("メール送信に失敗（データ更新は継続します）:", e)

# ---- 取得値の正規化 ----
def _norm_date(s):
    if not s: return ""
    s = re.sub(r"[年/]", "-", str(s)).replace("月", "-").replace("日", "")
    m = re.search(r"(\d{4})-(\d{1,2})-(\d{1,2})", s)
    return "%s-%02d-%02d" % (m.group(1), int(m.group(2)), int(m.group(3))) if m else ""
def _norm_yen(s):
    if s is None: return None
    t = re.sub(r"[^\d]", "", str(s));  return int(t) if t else None
def _norm_num(s):
    m = re.search(r"\d+(\.\d+)?", str(s or ""));  return float(m.group()) if m else None

# =========================================================
#  メイン
# =========================================================
def main(sample=False):
    log("=== START (sample=%s) ===" % sample)

    # 1. 取得
    new_cases = fetch_results(sample=sample)

    # 2. P2 参加者（detail_url がある案件のみ）
    if not sample:
        for c in new_cases:
            if c.get("detail_url"):
                c["bidders"] = fetch_bidders(c)

    # 3. P3 補強
    enrich(new_cases)

    # 4. 蓄積
    history = load_json(HISTORY, {"cases": []}).get("cases", [])
    merged, added = merge_history(history, new_cases)
    save_json(HISTORY, {"updated": TODAY, "cases": merged})
    log("蓄積:", len(merged), "件（新規", len(added), "件）")

    # 5. ツール用 data.json（直近2年に絞ると軽い。全件にしたければ merged をそのまま）
    cutoff = (datetime.date.today() - datetime.timedelta(days=730)).isoformat()
    recent = [c for c in merged if (c.get("date") or "") >= cutoff] or merged
    save_json(DATA_JSON, {"updated": TODAY, "cases": recent})
    log("data.json 出力:", len(recent), "件")

    # 6. 通知用：① 落札くんから書き出されたsales-status.csvで「未着手」のみ通す
    #            ② target_score で優先順位を算出、上位N件をメール
    state = load_json(STATE, {"notified": []})
    notified = set(state["notified"])
    status_map = _load_sales_status()
    if status_map:
        log("sales-status.csv 読込:", len(status_map), "社のステータスを認識")

    # 同社の累計落札件数（merged全体から集計）→ ターゲット度ボーナスに使う
    counts = _company_lookup_counts(merged)

    if sample:
        # sample モード：基本フィルタにかかった案件を全部通知対象に（テスト動作確認）
        pool = [c for c in merged if matches_target(c)]
        log("サンプル: 通知候補", len(pool), "件")
    else:
        # 本番：① 新規 ② 前回未通知 ③ 基本フィルタ合致
        pool = [c for c in added
                if matches_target(c) and c["id"] not in notified]

    # 営業ステータスで除外（連絡済・商談中・契約・失注・対象外などはメール対象外）
    pre_n = len(pool)
    pool = [c for c in pool if _is_actionable(c, status_map)]
    excluded = pre_n - len(pool)
    if excluded:
        log("営業ステータスで除外:", excluded, "件（手紙/FAX送付済・商談中・契約等）")

    # ターゲット度を計算 → 各案件にスコアと表示用情報を付与
    for c in pool:
        sc = target_score(c, counts)
        c["_score"] = sc
        c["_stars"] = _stars(sc)
        c["_company_count"] = counts.get(_normalize_co(c.get("company") or ""), 1)

    # スコア降順で並べ、上位N件をメール対象に
    pool.sort(key=lambda c: (-(c.get("_score") or 0),
                              -(c.get("amount") or 0),
                              0 if (c.get("pref") in KANTO_PREFS) else 1))
    top = pool[:MAIL_TOP_N]
    log("ターゲット度トップ%d社（プール %d社中）" % (len(top), len(pool)))

    send_email(top, len(added), status_excluded_n=excluded)

    if not sample:
        # 「新規 かつ 基本フィルタ合致」した案件は全てnotified入り。
        # （メール上位N件以外も再度同じケースが浮上しないよう、当日扱いに）
        # ただし営業ステータスで除外された会社は、後日ステータスが「未着手」に戻れば
        # 別の新規案件のときに再評価される（caseId単位なので別IDなら再通知される）。
        state["notified"] = list(notified | {c["id"] for c in added if matches_target(c)})
        save_json(STATE, state)

    log("=== DONE ===")

# =========================================================
#  ダミーデータ（--sample 用：全工程を通して動作確認するため）
# =========================================================
def _sample_cases():
    base = datetime.date.today()
    cos = [("東関建設(株)","千葉県",58),("房総土木(株)","千葉県",34),("関東道路(株)","神奈川県",72),
           ("常陸建設(株)","茨城県",41),("多摩川組(株)","東京都",63),("上総建設(株)","千葉県",88)]
    orgs = ["大宮国道事務所","横浜国道事務所","千葉国道事務所","常陸河川国道事務所"]
    out = []
    for i in range(10):
        w = cos[i % len(cos)]
        amt = (80 + i*18) * 1_000_000
        d = (base - datetime.timedelta(days=i*7)).isoformat()
        bidders = [{"company": w[0], "pref": w[1], "employees": w[2], "amount": amt}]
        for j in range(2):
            b = cos[(i+j+1) % len(cos)]
            bidders.append({"company": b[0], "pref": b[1], "employees": b[2], "amount": int(amt*1.05)})
        out.append({
            "date": d, "name": "国道%d号 ○○%s工事" % (16+i, ["道路改良","河川改修","舗装"][i%3]),
            "org": orgs[i % len(orgs)], "company": w[0], "pref": w[1], "employees": w[2],
            "amount": amt, "rate": 90.0 + (i % 9), "kind": classify_kind("道路改良"),
            "period": d + "〜" + (base + datetime.timedelta(days=200)).isoformat(),
            "docs": [{"label": "入札調書", "url": "https://www.ktr.mlit.go.jp/"}], "bidders": bidders, "detail_url": "",
        })
    # 業務（測量・コンサル等）も混在させる → 工事フィルタで除外されるのを確認できる
    for k, gname in enumerate(["一般国道○○号 橋梁点検業務", "△△川 河川測量業務",
                               "□□地区 道路詳細設計業務", "管内 用地補償コンサルタント業務"]):
        w = cos[k % len(cos)]
        r = _blank_rec()
        r.update({"date": (base - datetime.timedelta(days=k*5)).isoformat(), "name": gname,
                  "org": orgs[k % len(orgs)], "company": w[0], "pref": w[1], "employees": w[2],
                  "amount": (30 + k*12) * 1_000_000, "rate": 92.0 + k, "category": "業務"})
        out.append(r)
    return out


if __name__ == "__main__":
    main(sample=("--sample" in sys.argv))
